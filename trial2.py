import pandas as pd
from gpt4all import GPT4All
import json
import re
from openpyxl import load_workbook

model = GPT4All(model_name='Meta-Llama-3-8B-Instruct.Q4_0.gguf', model_path='./', allow_download=False, n_ctx=8192)

def read_excel(file_path):
    return pd.read_excel(file_path)

def prepare_prompt(df, num_samples=5, num_rows_to_generate=10):
    columns = df.columns.tolist()
    
    samples = df.sample(n=min(num_samples, len(df))).to_dict(orient='records')
    
    prompt = f"""
                You are a data generator. I will provide you with the column names of a dataset and a few sample rows. 
                Your task is to generate {num_rows_to_generate} new rows of synthetic data that follow the same structure and pattern.

                Please be creative with the generated data. Make sure that they are unique from each other.

                Column Names: {columns}

                Sample Rows:
                {json.dumps(samples, indent=2)}

                Generate {num_rows_to_generate} rows of synthetic data in JSON format. 
                Make sure that the JSON should start and end with square brackets, i.e., []
            """
    return prompt

def extract_json(response):
    try:
        # Look for the first [ and last ] in the response
        json_str = re.search(r"\[.*\]", response, re.DOTALL).group(0)
        return json.loads(json_str)
    except (AttributeError, json.JSONDecodeError) as e:
        print(f"Error extracting JSON: {e}")
        return None

def generate_synthetic_data(df, num_samples=2, num_rows=10):
    if num_rows < 2:
        num_rows = 2

    prompt = prepare_prompt(df, num_samples=num_samples, num_rows_to_generate=num_rows)
    
    with model.chat_session():
        response = model.generate(prompt, max_tokens = num_rows * 1024)

        print(response)

        try:
            synthetic_data = extract_json(response)
            return pd.DataFrame(synthetic_data)
        except json.JSONDecodeError:
            print("Failed to parse JSON from LLM response.")
            return None
        
def append_to_excel(file_path, sheet_name, df):
    try:
        book = load_workbook(file_path)
        if sheet_name in book.sheetnames:
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                existing_data = pd.read_excel(file_path, sheet_name=sheet_name)
                combined_df = pd.concat([existing_data, df], ignore_index=True)
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
    except FileNotFoundError:
        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def main():
    file_path = "Dataset.xlsx"
    df = read_excel(file_path)

    ## Iss line pe num_samples is ki kitne datapoints uthayenge original dataset se
    ## And num_rows is kitne synthetic datapoints generate krenge
    synthetic_df = generate_synthetic_data(df, num_samples=8, num_rows=8)

    if synthetic_df is not None:
        print("Synthetic DataFrame:")
        print(synthetic_df)
        append_to_excel(file_path, "Synthetic Data", synthetic_df)
    else:
        print("Failed to generate synthetic data.")

if __name__ == "__main__":
    main()
